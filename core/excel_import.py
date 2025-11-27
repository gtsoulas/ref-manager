import openpyxl
from django.contrib.auth.models import User
from django.db import transaction
from decimal import Decimal
from datetime import datetime
from .models import Colleague, Output, CriticalFriend


class ExcelImporter:
    """Import data from Excel files into the database"""
    
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.imported_count = 0
    
    def import_colleagues(self, file_path):
        """Import colleagues from Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            headers = [cell.value for cell in sheet[1]]
            
            with transaction.atomic():
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        data = dict(zip(headers, row))
                        
                        if not data.get('Staff ID'):
                            continue
                        
                        username = data.get('Email', '').split('@')[0] or f"user_{data['Staff ID']}"
                        user, created = User.objects.get_or_create(
                            username=username,
                            defaults={
                                'first_name': data.get('First Name', ''),
                                'last_name': data.get('Last Name', ''),
                                'email': data.get('Email', ''),
                            }
                        )
                        
                        fte = float(data.get('FTE', 1.0))
                        if fte < 0.1 or fte > 1.0:
                            self.warnings.append(f"Row {row_num}: FTE out of range, using 1.0")
                            fte = 1.0
                        
                        contract_map = {
                            'permanent': 'permanent',
                            'fixed-term': 'fixed-term',
                            'fixed term': 'fixed-term',
                            'research': 'research',
                        }
                        contract_type = contract_map.get(
                            str(data.get('Contract Type', 'permanent')).lower(),
                            'permanent'
                        )
                        
                        returnable_str = str(data.get('Is Returnable', 'yes')).lower()
                        is_returnable = returnable_str in ['yes', 'y', 'true', '1']
                        
                        colleague, created = Colleague.objects.update_or_create(
                            staff_id=data['Staff ID'],
                            defaults={
                                'user': user,
                                'title': data.get('Title', ''),
                                'fte': Decimal(str(fte)),
                                'contract_type': contract_type,
                                'unit_of_assessment': data.get('Unit of Assessment', 'Unknown'),
                                'is_returnable': is_returnable,
                            }
                        )
                        
                        self.imported_count += 1
                        
                    except Exception as e:
                        self.errors.append(f"Row {row_num}: {str(e)}")
            
            return True
            
        except Exception as e:
            self.errors.append(f"Error reading file: {str(e)}")
            return False
    
    def import_outputs(self, file_path):
        """Import outputs from Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            headers = [cell.value for cell in sheet[1]]
            
            with transaction.atomic():
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        data = dict(zip(headers, row))
                        
                        if not data.get('Staff ID') or not data.get('Title'):
                            continue
                        
                        try:
                            colleague = Colleague.objects.get(staff_id=data['Staff ID'])
                        except Colleague.DoesNotExist:
                            self.errors.append(f"Row {row_num}: Staff ID '{data['Staff ID']}' not found")
                            continue
                        
                        pub_date = self._parse_date(data.get('Publication Date'))
                        if not pub_date:
                            pub_date = datetime.now().date()
                            self.warnings.append(f"Row {row_num}: Invalid date, using today")
                        
                        quality_map = {
                            '4*': '4*', '4': '4*',
                            '3*': '3*', '3': '3*',
                            '2*': '2*', '2': '2*',
                            '1*': '1*', '1': '1*',
                            'u': 'U', 'unclassified': 'U',
                        }
                        quality_rating = quality_map.get(
                            str(data.get('Quality Rating', 'U')).lower(),
                            'U'
                        )
                        
                        is_open_access = str(data.get('Is Open Access', 'no')).lower() in ['yes', 'y', 'true', '1']
                        is_double_weighted = str(data.get('Is Double Weighted', 'no')).lower() in ['yes', 'y', 'true', '1']
                        is_interdisciplinary = str(data.get('Is Interdisciplinary', 'no')).lower() in ['yes', 'y', 'true', '1']
                        
                        output = Output.objects.create(
                            colleague=colleague,
                            title=data['Title'],
                            all_authors=data.get('All Authors', ''),
                            author_position=int(data.get('Author Position', 1)),
                            publication_type=data.get('Publication Type', 'A'),
                            publication_date=pub_date,
                            publication_venue=data.get('Publication Venue', ''),
                            quality_rating=quality_rating,
                            doi=data.get('DOI', ''),
                            url=data.get('URL', ''),
                            abstract=data.get('Abstract', ''),
                            is_open_access=is_open_access,
                            is_double_weighted=is_double_weighted,
                            is_interdisciplinary=is_interdisciplinary,
                            status='draft',
                        )
                        
                        self.imported_count += 1
                        
                    except Exception as e:
                        self.errors.append(f"Row {row_num}: {str(e)}")
            
            return True
            
        except Exception as e:
            self.errors.append(f"Error reading file: {str(e)}")
            return False
    
    def import_critical_friends(self, file_path):
        """Import critical friends from Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            headers = [cell.value for cell in sheet[1]]
            
            with transaction.atomic():
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        data = dict(zip(headers, row))
                        
                        if not data.get('Name') or not data.get('Email'):
                            continue
                        
                        cf, created = CriticalFriend.objects.update_or_create(
                            email=data['Email'],
                            defaults={
                                'name': data['Name'],
                                'institution': data.get('Institution', ''),
                                'expertise_area': data.get('Expertise Area', ''),
                                'bio': data.get('Bio', ''),
                            }
                        )
                        
                        self.imported_count += 1
                        
                    except Exception as e:
                        self.errors.append(f"Row {row_num}: {str(e)}")
            
            return True
            
        except Exception as e:
            self.errors.append(f"Error reading file: {str(e)}")
            return False
    
    def _parse_date(self, date_value):
        """Parse various date formats"""
        if not date_value:
            return None
        
        if isinstance(date_value, datetime):
            return date_value.date()
        
        formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']
        for fmt in formats:
            try:
                return datetime.strptime(str(date_value), fmt).date()
            except:
                continue
        
        return None
