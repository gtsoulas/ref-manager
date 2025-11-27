# core/views_export.py - CORRECTED VERSION with Both Internal & External Reviewers
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import csv
from datetime import datetime
from .models import (
    CriticalFriendAssignment, 
    InternalPanelAssignment,
    CriticalFriend, 
    InternalPanelMember,
    Colleague
)

@login_required
def export_assignments_view(request):
    """Display export options and filters"""
    # Get both critical friends and internal panel members
    critical_friends = CriticalFriend.objects.all().order_by('name')
    internal_panel = InternalPanelMember.objects.select_related('colleague__user').all().order_by('colleague__user__last_name')
    
    colleagues = Colleague.objects.all().order_by('user__last_name', 'user__first_name')
    
    # Get status choices from both models
    cf_status_choices = CriticalFriendAssignment.STATUS_CHOICES
    ip_status_choices = InternalPanelAssignment.STATUS_CHOICES if hasattr(InternalPanelAssignment, 'STATUS_CHOICES') else []
    
    # Count assignments
    cf_count = CriticalFriendAssignment.objects.count()
    ip_count = InternalPanelAssignment.objects.count()
    
    context = {
        'critical_friends': critical_friends,
        'internal_panel': internal_panel,
        'authors': colleagues,
        'status_choices': cf_status_choices,
        'cf_count': cf_count,
        'ip_count': ip_count,
    }
    return render(request, 'core/export_assignments.html', context)

@login_required
def export_assignments_excel(request):
    """Export assignments to Excel with file and DOI links - handles both internal and external"""
    # Get filter parameters
    recipient_type = request.GET.get('recipient_type', 'all')
    reviewer_param = request.GET.get('reviewer', '')  # This will be like 'cf_1' or 'ip_2'
    author_id = request.GET.get('author')
    status = request.GET.get('status')
    
    # Parse the reviewer parameter to determine type and ID
    reviewer_type = None  # 'cf' for critical friend, 'ip' for internal panel
    reviewer_id = None
    
    if reviewer_param:
        if reviewer_param.startswith('cf_'):
            reviewer_type = 'cf'
            reviewer_id = reviewer_param.replace('cf_', '')
        elif reviewer_param.startswith('ip_'):
            reviewer_type = 'ip'
            reviewer_id = reviewer_param.replace('ip_', '')
    
    # Collect assignments from both systems
    assignments = []
    
    # Get Critical Friend assignments if needed
    if recipient_type in ['all', 'critical_friends']:
        cf_assignments = CriticalFriendAssignment.objects.select_related(
            'output', 'critical_friend', 'output__colleague', 'output__colleague__user'
        ).all()
        
        # Apply filters for critical friends
        if reviewer_type == 'cf' and reviewer_id:
            cf_assignments = cf_assignments.filter(critical_friend_id=reviewer_id)
        
        if author_id:
            cf_assignments = cf_assignments.filter(output__colleague_id=author_id)
        
        if status:
            cf_assignments = cf_assignments.filter(status=status)
        
        # Add to assignments list with type marker
        for assignment in cf_assignments:
            assignments.append({
                'type': 'critical_friend',
                'assignment': assignment
            })
    
    # Get Internal Panel assignments if needed
    if recipient_type in ['all', 'internal']:
        ip_assignments = InternalPanelAssignment.objects.select_related(
            'output', 'panel_member', 'panel_member__colleague', 'panel_member__colleague__user',
            'output__colleague', 'output__colleague__user'
        ).all()
        
        # Apply filters for internal panel
        if reviewer_type == 'ip' and reviewer_id:
            ip_assignments = ip_assignments.filter(panel_member_id=reviewer_id)
        
        if author_id:
            ip_assignments = ip_assignments.filter(output__colleague_id=author_id)
        
        if status:
            ip_assignments = ip_assignments.filter(status=status)
        
        # Add to assignments list with type marker
        for assignment in ip_assignments:
            assignments.append({
                'type': 'internal_panel',
                'assignment': assignment
            })
    
    # Sort combined assignments by reviewer name, then output title
    assignments.sort(key=lambda x: (
        x['assignment'].critical_friend.name if x['type'] == 'critical_friend' 
        else x['assignment'].panel_member.colleague.user.get_full_name(),
        x['assignment'].output.title
    ))
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review Assignments"
    
    # Define headers
    headers = [
        'Reviewer Type',
        'Reviewer Name',
        'Reviewer Email',
        'Output Title',
        'Author',
        'Output Type',
        'File Link',
        'DOI Link',
        'Status',
        'Due Date',
        'Quality Rating',
        'Review Notes',
        'Special Instructions'
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Color fills for different reviewer types
    internal_fill = PatternFill(start_color='D6E9F8', end_color='D6E9F8', fill_type='solid')  # Light blue
    external_fill = PatternFill(start_color='FFF4CC', end_color='FFF4CC', fill_type='solid')  # Light yellow
    
    # Write data
    for row_num, item in enumerate(assignments, 2):
        assignment = item['assignment']
        is_internal = item['type'] == 'internal_panel'
        output = assignment.output
        
        # Apply row color based on type
        row_fill = internal_fill if is_internal else external_fill
        
        # Reviewer Type
        cell = ws.cell(row=row_num, column=1, value='Internal Panel' if is_internal else 'Critical Friend')
        cell.fill = row_fill
        
        # Reviewer Name
        if is_internal:
            reviewer_name = assignment.panel_member.colleague.user.get_full_name()
            reviewer_email = assignment.panel_member.colleague.user.email
        else:
            reviewer_name = assignment.critical_friend.name
            reviewer_email = assignment.critical_friend.email
        
        cell = ws.cell(row=row_num, column=2, value=reviewer_name)
        cell.fill = row_fill
        
        # Reviewer Email
        cell = ws.cell(row=row_num, column=3, value=reviewer_email)
        cell.fill = row_fill
        
        # Output Title
        cell = ws.cell(row=row_num, column=4, value=output.title)
        cell.fill = row_fill
        
        # Author name
        author_name = output.colleague.user.get_full_name()
        cell = ws.cell(row=row_num, column=5, value=author_name)
        cell.fill = row_fill
        
        # Output type
        cell = ws.cell(row=row_num, column=6, value=output.get_publication_type_display())
        cell.fill = row_fill
        
        # FILE LINK (Column 7)
        file_cell = ws.cell(row=row_num, column=7)
        file_cell.fill = row_fill
        if output.pdf_file:
            try:
                file_url = request.build_absolute_uri(output.pdf_file.url)
                file_cell.value = file_url
                file_cell.hyperlink = file_url
                file_cell.font = Font(color='0563C1', underline='single')
            except:
                file_cell.value = 'File available'
        elif output.url:
            file_cell.value = output.url
            file_cell.hyperlink = output.url
            file_cell.font = Font(color='0563C1', underline='single')
        else:
            file_cell.value = 'No file'
        
        # DOI LINK (Column 8)
        doi_cell = ws.cell(row=row_num, column=8)
        doi_cell.fill = row_fill
        if output.doi:
            doi = output.doi.strip()
            if doi.startswith('http'):
                doi_url = doi
            elif doi.startswith('10.'):
                doi_url = f'https://doi.org/{doi}'
            else:
                doi_url = f'https://doi.org/{doi}'
            
            doi_cell.value = doi
            doi_cell.hyperlink = doi_url
            doi_cell.font = Font(color='0563C1', underline='single')
        elif output.url and not output.pdf_file:
            doi_cell.value = output.url
            doi_cell.hyperlink = output.url
            doi_cell.font = Font(color='0563C1', underline='single')
        else:
            doi_cell.value = 'No DOI'
        
        # Status
        cell = ws.cell(row=row_num, column=9, value=assignment.get_status_display())
        cell.fill = row_fill
        
        # Due date
        due_date = ''
        if hasattr(assignment, 'due_date') and assignment.due_date:
            due_date = assignment.due_date.strftime('%Y-%m-%d')
        elif hasattr(assignment, 'review_date') and assignment.review_date:
            due_date = assignment.review_date.strftime('%Y-%m-%d')
        else:
            due_date = 'Not set'
        cell = ws.cell(row=row_num, column=10, value=due_date)
        cell.fill = row_fill
        
        # Quality rating
        quality = ''
        if hasattr(assignment, 'quality_rating') and assignment.quality_rating:
            quality = assignment.quality_rating
        elif hasattr(assignment, 'rating_recommendation') and assignment.rating_recommendation:
            quality = assignment.rating_recommendation
        elif output.quality_rating:
            quality = output.quality_rating
        cell = ws.cell(row=row_num, column=11, value=quality)
        cell.fill = row_fill
        
        # Review notes
        notes = ''
        if hasattr(assignment, 'review_notes') and assignment.review_notes:
            notes = assignment.review_notes
        elif hasattr(assignment, 'comments') and assignment.comments:
            notes = assignment.comments
        cell = ws.cell(row=row_num, column=12, value=notes)
        cell.fill = row_fill
        
        # Special instructions
        instructions = ''
        if hasattr(assignment, 'special_instructions') and assignment.special_instructions:
            instructions = assignment.special_instructions
        elif hasattr(assignment, 'notes') and assignment.notes:
            instructions = assignment.notes
        cell = ws.cell(row=row_num, column=13, value=instructions)
        cell.fill = row_fill
    
    # Adjust column widths
    column_widths = {
        'A': 15,  # Reviewer Type
        'B': 20,  # Reviewer Name
        'C': 30,  # Reviewer Email
        'D': 50,  # Output Title
        'E': 20,  # Author
        'F': 15,  # Output Type
        'G': 40,  # File Link
        'H': 40,  # DOI Link
        'I': 15,  # Status
        'J': 12,  # Due Date
        'K': 15,  # Quality Rating
        'L': 40,  # Review Notes
        'M': 40,  # Special Instructions
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Prepare response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ref_assignments_with_links_{timestamp}.xlsx'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def export_assignments_csv(request):
    """Export assignments to CSV - handles both internal and external"""
    # Get filter parameters
    recipient_type = request.GET.get('recipient_type', 'all')
    reviewer_param = request.GET.get('reviewer', '')
    author_id = request.GET.get('author')
    status = request.GET.get('status')
    
    # Parse the reviewer parameter
    reviewer_type = None
    reviewer_id = None
    
    if reviewer_param:
        if reviewer_param.startswith('cf_'):
            reviewer_type = 'cf'
            reviewer_id = reviewer_param.replace('cf_', '')
        elif reviewer_param.startswith('ip_'):
            reviewer_type = 'ip'
            reviewer_id = reviewer_param.replace('ip_', '')
    
    # Collect assignments from both systems
    assignments = []
    
    # Get Critical Friend assignments
    if recipient_type in ['all', 'critical_friends']:
        cf_assignments = CriticalFriendAssignment.objects.select_related(
            'output', 'critical_friend', 'output__colleague', 'output__colleague__user'
        ).all()
        
        if reviewer_type == 'cf' and reviewer_id:
            cf_assignments = cf_assignments.filter(critical_friend_id=reviewer_id)
        if author_id:
            cf_assignments = cf_assignments.filter(output__colleague_id=author_id)
        if status:
            cf_assignments = cf_assignments.filter(status=status)
        
        for assignment in cf_assignments:
            assignments.append({'type': 'critical_friend', 'assignment': assignment})
    
    # Get Internal Panel assignments
    if recipient_type in ['all', 'internal']:
        ip_assignments = InternalPanelAssignment.objects.select_related(
            'output', 'panel_member', 'panel_member__colleague', 'panel_member__colleague__user',
            'output__colleague', 'output__colleague__user'
        ).all()
        
        if reviewer_type == 'ip' and reviewer_id:
            ip_assignments = ip_assignments.filter(panel_member_id=reviewer_id)
        if author_id:
            ip_assignments = ip_assignments.filter(output__colleague_id=author_id)
        if status:
            ip_assignments = ip_assignments.filter(status=status)
        
        for assignment in ip_assignments:
            assignments.append({'type': 'internal_panel', 'assignment': assignment})
    
    # Sort combined assignments
    assignments.sort(key=lambda x: (
        x['assignment'].critical_friend.name if x['type'] == 'critical_friend' 
        else x['assignment'].panel_member.colleague.user.get_full_name(),
        x['assignment'].output.title
    ))
    
    # Create CSV response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ref_assignments_with_links_{timestamp}.csv'
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow([
        'Reviewer Type',
        'Reviewer Name',
        'Reviewer Email',
        'Output Title',
        'Author',
        'Output Type',
        'File Link',
        'DOI Link',
        'Status',
        'Due Date',
        'Quality Rating',
        'Review Notes',
        'Special Instructions'
    ])
    
    # Write data
    for item in assignments:
        assignment = item['assignment']
        is_internal = item['type'] == 'internal_panel'
        output = assignment.output
        
        # Reviewer info
        if is_internal:
            reviewer_name = assignment.panel_member.colleague.user.get_full_name()
            reviewer_email = assignment.panel_member.colleague.user.email
        else:
            reviewer_name = assignment.critical_friend.name
            reviewer_email = assignment.critical_friend.email
        
        # File link
        file_link = ''
        if output.pdf_file:
            try:
                file_link = request.build_absolute_uri(output.pdf_file.url)
            except:
                file_link = 'File available'
        elif output.url:
            file_link = output.url
        else:
            file_link = 'No file'
        
        # DOI link
        doi_link = ''
        if output.doi:
            doi = output.doi.strip()
            if doi.startswith('http'):
                doi_link = doi
            elif doi.startswith('10.'):
                doi_link = f'https://doi.org/{doi}'
            else:
                doi_link = f'https://doi.org/{doi}'
        elif output.url and not output.pdf_file:
            doi_link = output.url
        else:
            doi_link = 'No DOI'
        
        # Due date
        due_date = ''
        if hasattr(assignment, 'due_date') and assignment.due_date:
            due_date = assignment.due_date.strftime('%Y-%m-%d')
        elif hasattr(assignment, 'review_date') and assignment.review_date:
            due_date = assignment.review_date.strftime('%Y-%m-%d')
        else:
            due_date = 'Not set'
        
        # Quality rating
        quality = ''
        if hasattr(assignment, 'quality_rating') and assignment.quality_rating:
            quality = assignment.quality_rating
        elif hasattr(assignment, 'rating_recommendation') and assignment.rating_recommendation:
            quality = assignment.rating_recommendation
        elif output.quality_rating:
            quality = output.quality_rating
        
        # Notes
        notes = ''
        if hasattr(assignment, 'review_notes') and assignment.review_notes:
            notes = assignment.review_notes
        elif hasattr(assignment, 'comments') and assignment.comments:
            notes = assignment.comments
        
        # Instructions
        instructions = ''
        if hasattr(assignment, 'special_instructions') and assignment.special_instructions:
            instructions = assignment.special_instructions
        elif hasattr(assignment, 'notes') and assignment.notes:
            instructions = assignment.notes
        
        author_name = output.colleague.user.get_full_name()
        
        writer.writerow([
            'Internal Panel' if is_internal else 'Critical Friend',
            reviewer_name,
            reviewer_email,
            output.title,
            author_name,
            output.get_publication_type_display(),
            file_link,
            doi_link,
            assignment.get_status_display(),
            due_date,
            quality,
            notes,
            instructions
        ])
    
    return response
