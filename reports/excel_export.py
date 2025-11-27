# ============================================================
# FILE: reports/excel_export.py
# Excel export functionality for risk analysis reports
# ============================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from decimal import Decimal


def export_risk_analysis_to_excel(outputs=None, submission=None):
    """
    Export comprehensive risk analysis to Excel workbook.
    
    Args:
        outputs: QuerySet of Output objects (if None, exports all)
        submission: REFSubmission object (optional, for submission-specific export)
    
    Returns:
        HttpResponse with Excel file
    """
    from core.models import Output, REFSubmission
    
    if outputs is None:
        outputs = Output.objects.all()
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    create_summary_sheet(wb, outputs, submission)
    create_outputs_detail_sheet(wb, outputs)
    create_risk_matrix_sheet(wb, outputs)
    
    if submission:
        create_submission_analysis_sheet(wb, submission)
    
    # Prepare HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'REF_Risk_Analysis_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    if submission:
        filename = f'REF_{submission.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def create_summary_sheet(wb, outputs, submission=None):
    """Create summary overview sheet"""
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = 'REF Risk Analysis Summary'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    
    if submission:
        ws['A3'] = f'Submission: {submission.name}'
        ws['A4'] = f'UOA: {submission.uoa}'
    
    # Overall Statistics
    row = 6
    ws[f'A{row}'] = 'Overall Statistics'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    row += 1
    stats = [
        ('Total Outputs', outputs.count()),
        ('Average Risk Score', f"{outputs.aggregate(models.Avg('overall_risk_score'))['overall_risk_score__avg'] or 0:.2f}"),
        ('Average Quality', f"{sum(o.get_quality_value() for o in outputs) / outputs.count() if outputs.count() > 0 else 0:.2f}*"),
    ]
    
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Risk Distribution
    row += 1
    ws[f'A{row}'] = 'Risk Distribution'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    row += 1
    ws[f'A{row}'] = 'Risk Level'
    ws[f'B{row}'] = 'Count'
    ws[f'C{row}'] = 'Percentage'
    
    # Header styling
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
    
    row += 1
    total = outputs.count()
    
    risk_levels = [
        ('Low (<0.25)', outputs.filter(overall_risk_score__lt=Decimal('0.25')).count(), '28a745'),
        ('Medium-Low (0.25-0.50)', outputs.filter(
            overall_risk_score__gte=Decimal('0.25'),
            overall_risk_score__lt=Decimal('0.50')
        ).count(), 'ffc107'),
        ('Medium-High (0.50-0.75)', outputs.filter(
            overall_risk_score__gte=Decimal('0.50'),
            overall_risk_score__lt=Decimal('0.75')
        ).count(), 'fd7e14'),
        ('High (≥0.75)', outputs.filter(overall_risk_score__gte=Decimal('0.75')).count(), 'dc3545'),
    ]
    
    for level, count, color in risk_levels:
        ws[f'A{row}'] = level
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{count/total*100:.1f}%" if total > 0 else "0%"
        
        # Color coding
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        if color in ['28a745', 'dc3545', 'fd7e14']:
            ws[f'A{row}'].font = Font(color='FFFFFF')
        
        row += 1
    
    # Quality Distribution
    row += 1
    ws[f'A{row}'] = 'Quality Distribution'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    row += 1
    ws[f'A{row}'] = 'Quality Rating'
    ws[f'B{row}'] = 'Count'
    
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
    
    row += 1
    quality_counts = [
        ('4*', outputs.filter(quality_rating='4*').count()),
        ('3*', outputs.filter(quality_rating='3*').count()),
        ('2*', outputs.filter(quality_rating='2*').count()),
        ('1*', outputs.filter(quality_rating='1*').count()),
        ('Unclassified', outputs.filter(quality_rating='unclassified').count()),
    ]
    
    for rating, count in quality_counts:
        ws[f'A{row}'] = rating
        ws[f'B{row}'] = count
        row += 1
    
    # Auto-size columns
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def create_outputs_detail_sheet(wb, outputs):
    """Create detailed outputs list with risk scores"""
    ws = wb.create_sheet("Outputs Detail")
    
    # Headers
    headers = [
        'ID', 'Title', 'Colleague', 'Quality', 'Publication Status',
        'Overall Risk', 'Content Risk', 'Timeline Risk', 'Risk Level',
        'OA Compliance Risk', 'Panel Alignment', 'Venue Prestige',
        'REF Ready'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row, output in enumerate(outputs, 2):
        ws.cell(row=row, column=1, value=output.id)
        ws.cell(row=row, column=2, value=output.title)
        ws.cell(row=row, column=3, value=str(output.colleague) if hasattr(output, 'colleague') else '')
        ws.cell(row=row, column=4, value=output.quality_rating)
        ws.cell(row=row, column=5, value=output.publication_status)
        ws.cell(row=row, column=6, value=float(output.overall_risk_score))
        ws.cell(row=row, column=7, value=float(output.content_risk_score))
        ws.cell(row=row, column=8, value=float(output.timeline_risk_score))
        ws.cell(row=row, column=9, value=output.get_risk_level_display())
        ws.cell(row=row, column=10, value='Yes' if output.oa_compliance_risk else 'No')
        ws.cell(row=row, column=11, value=float(output.panel_alignment_score))
        ws.cell(row=row, column=12, value=float(output.venue_prestige_score))
        ws.cell(row=row, column=13, value='Yes' if output.is_ref_ready() else 'No')
        
        # Color code risk level
        risk_level = output.get_risk_level()
        risk_colors = {
            'low': '28a745',
            'medium-low': 'ffc107',
            'medium-high': 'fd7e14',
            'high': 'dc3545'
        }
        
        if risk_level in risk_colors:
            for col in [6, 9]:  # Overall Risk and Risk Level columns
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=risk_colors[risk_level],
                    end_color=risk_colors[risk_level],
                    fill_type='solid'
                )
                if risk_level in ['low', 'medium-high', 'high']:
                    ws.cell(row=row, column=col).font = Font(color='FFFFFF')
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws.column_dimensions['B'].width = 50  # Title column wider
    
    # Freeze header row
    ws.freeze_panes = 'A2'


def create_risk_matrix_sheet(wb, outputs):
    """Create quality vs risk matrix"""
    ws = wb.create_sheet("Risk Matrix")
    
    ws['A1'] = 'Quality vs Risk Matrix'
    ws['A1'].font = Font(size=14, bold=True)
    
    # Headers
    ws['A3'] = 'Output'
    ws['B3'] = 'Quality Value'
    ws['C3'] = 'Risk Score'
    
    for col in ['A', 'B', 'C']:
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws[f'{col}3'].font = Font(bold=True, color='FFFFFF')
    
    # Data
    row = 4
    for output in outputs:
        ws[f'A{row}'] = output.title[:50]
        ws[f'B{row}'] = output.get_quality_value()
        ws[f'C{row}'] = float(output.overall_risk_score)
        row += 1
    
    # Create scatter chart (simulated as bar chart due to openpyxl limitations)
    # Note: openpyxl doesn't support scatter charts well, so we create a bar chart
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def create_submission_analysis_sheet(wb, submission):
    """Create detailed submission analysis sheet"""
    ws = wb.create_sheet("Submission Analysis")
    
    # Title
    ws['A1'] = f'Submission: {submission.name}'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f'UOA: {submission.uoa}'
    ws['A3'] = f'Year: {submission.submission_year}'
    
    # Portfolio Metrics
    row = 5
    ws[f'A{row}'] = 'Portfolio Metrics'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    row += 1
    metrics = [
        ('Overall Portfolio Score', f"{submission.get_overall_portfolio_score():.2f} / 4.00"),
        ('Quality Score', f"{submission.portfolio_quality_score:.2f}"),
        ('Risk Score (inverted)', f"{submission.portfolio_risk_score:.2f}"),
        ('Representativeness', f"{submission.representativeness_score:.2f}"),
        ('Staff Inclusion', f"{submission.equality_score:.1f}%"),
        ('Gender Balance', f"{submission.gender_balance_score:.2f}"),
    ]
    
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Weights
    row += 1
    ws[f'A{row}'] = 'Optimization Weights'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    row += 1
    weights = [
        ('Quality Weight', float(submission.weight_quality)),
        ('Risk Weight', float(submission.weight_risk)),
        ('Representativeness Weight', float(submission.weight_representativeness)),
        ('Equality Weight', float(submission.weight_equality)),
        ('Gender Balance Weight', float(submission.weight_gender_balance)),
    ]
    
    for label, value in weights:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Readiness Assessment
    row += 1
    readiness = submission.get_submission_readiness()
    
    ws[f'A{row}'] = 'Submission Readiness'
    ws[f'A{row}'].font = Font(size=14, bold=True)
    
    row += 1
    ws[f'A{row}'] = 'Status'
    ws[f'B{row}'] = 'READY' if readiness['ready'] else 'NOT READY'
    ws[f'B{row}'].font = Font(bold=True, color='008000' if readiness['ready'] else 'FF0000')
    
    row += 1
    ws[f'A{row}'] = 'Readiness Percentage'
    ws[f'B{row}'] = f"{readiness['readiness_percentage']:.1f}%"
    
    row += 1
    ws[f'A{row}'] = 'Ready Outputs'
    ws[f'B{row}'] = f"{readiness['ready_outputs']} / {readiness['total_outputs']}"
    
    if readiness['issues']:
        row += 1
        ws[f'A{row}'] = 'Issues'
        ws[f'A{row}'].font = Font(bold=True)
        for issue in readiness['issues']:
            row += 1
            ws[f'B{row}'] = f"• {issue}"
    
    # Auto-size
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30


# ============================================================
# USAGE IN VIEWS:
# ============================================================
#
# from reports.excel_export import export_risk_analysis_to_excel
#
# def export_risk_excel(request):
#     """View to export risk analysis as Excel"""
#     outputs = Output.objects.all()
#     return export_risk_analysis_to_excel(outputs)
#
# def export_submission_excel(request, pk):
#     """View to export submission risk profile as Excel"""
#     submission = get_object_or_404(REFSubmission, pk=pk)
#     outputs = submission.outputs.all()
#     return export_risk_analysis_to_excel(outputs, submission)
#
# # Add to urls.py:
# path('export/risk-analysis/', views.export_risk_excel, name='risk-export-excel'),
# path('submissions/<int:pk>/export/', views.export_submission_excel, name='submission-export-excel'),
#
